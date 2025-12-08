from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from app.db import get_session
from app import models
from app.schemas import GradeSettingIn, AssignmentOut, AdminSettingIn, AdminSettingOut
from app.assignment.engine import run_assignment
from app.core.security import get_current_user
from fastapi.responses import StreamingResponse
import io
from openpyxl import load_workbook
from typing import List, Dict, Any

router = APIRouter(prefix="/admin", tags=["admin"])


@router.get("/dashboard")
async def dashboard(
  year: int,
  session: AsyncSession = Depends(get_session),
  user=Depends(get_current_user),
):
  if user.get("role") != "admin":
    raise HTTPException(status_code=403, detail="admin only")
  
  # 전체 교사 수 (설정에서 가져오기, 없으면 DB의 실제 교사 수)
  admin_setting_stmt = select(models.AdminSetting).where(models.AdminSetting.year == year)
  admin_setting_res = await session.execute(admin_setting_stmt)
  admin_setting = admin_setting_res.scalars().first()
  total_count = admin_setting.total_teachers if admin_setting else 0
  
  # 설정이 없으면 DB의 실제 교사 수 사용
  if total_count == 0:
    total_teachers = (await session.execute(select(models.Teacher))).scalars().all()
    total_count = len(total_teachers)
  
  # 희망 제출 수
  prefs = (
    await session.execute(select(models.Preference).where(models.Preference.year == year))
  ).scalars().all()
  submitted_count = len(prefs)
  
  # 필요 담임 수
  settings = (
    await session.execute(select(models.GradeSetting).where(models.GradeSetting.year == year))
  ).scalars().all()
  required_homerooms = sum(s.required_homerooms for s in settings)
  
  # 학년별 학급 수
  grade_class_counts = {s.grade: s.class_count for s in settings}
  
  # 마감 여부
  is_closed = admin_setting.is_closed if admin_setting else False
  
  # 지망 현황
  first_counts = {}
  second_counts = {}
  third_counts = {}
  for p in prefs:
    if p.first_choice_grade:
      first_counts[p.first_choice_grade] = first_counts.get(p.first_choice_grade, 0) + 1
    if p.second_choice_grade:
      second_counts[p.second_choice_grade] = second_counts.get(p.second_choice_grade, 0) + 1
    if p.third_choice_grade:
      third_counts[p.third_choice_grade] = third_counts.get(p.third_choice_grade, 0) + 1
  
  return {
    "year": year,
    "total_teachers": total_count,
    "submitted_count": submitted_count,
    "required_homerooms": required_homerooms,
    "grade_class_counts": grade_class_counts,
    "first_choice_counts": first_counts,
    "second_choice_counts": second_counts,
    "third_choice_counts": third_counts,
  }


@router.get("/summary")
async def summary(
  year: int,
  session: AsyncSession = Depends(get_session),
  user=Depends(get_current_user),
):
  if user.get("role") != "admin":
    raise HTTPException(status_code=403, detail="admin only")
  prefs = (
    await session.execute(select(models.Preference).where(models.Preference.year == year))
  ).scalars()
  first_counts = {}
  second_counts = {}
  third_counts = {}
  for p in prefs:
    if p.first_choice_grade:
      first_counts[p.first_choice_grade] = first_counts.get(p.first_choice_grade, 0) + 1
    if p.second_choice_grade:
      second_counts[p.second_choice_grade] = second_counts.get(p.second_choice_grade, 0) + 1
    if p.third_choice_grade:
      third_counts[p.third_choice_grade] = third_counts.get(p.third_choice_grade, 0) + 1
  return {
    "year": year,
    "first_choice_counts": first_counts,
    "second_choice_counts": second_counts,
    "third_choice_counts": third_counts,
  }


@router.get("/settings")
async def get_settings(
  year: int,
  session: AsyncSession = Depends(get_session),
  user=Depends(get_current_user),
):
  if user.get("role") != "admin":
    raise HTTPException(status_code=403, detail="admin only")
  stmt = select(models.GradeSetting).where(models.GradeSetting.year == year)
  res = await session.execute(stmt)
  return res.scalars().all()


@router.post("/settings")
async def upsert_settings(
  payload: list[GradeSettingIn],
  session: AsyncSession = Depends(get_session),
  user=Depends(get_current_user),
):
  if user.get("role") != "admin":
    raise HTTPException(status_code=403, detail="admin only")
  for item in payload:
    stmt = select(models.GradeSetting).where(
      models.GradeSetting.year == item.year,
      models.GradeSetting.grade == item.grade,
    )
    res = await session.execute(stmt)
    gs = res.scalars().first()
    subject_teachers = item.required_subject_teachers if hasattr(item, "required_subject_teachers") else 0
    duty_heads = item.required_duty_heads if hasattr(item, "required_duty_heads") else 0
    if gs:
      gs.class_count = item.class_count
      gs.required_homerooms = item.required_homerooms
      gs.required_subject_teachers = subject_teachers
      gs.required_duty_heads = duty_heads
    else:
      gs = models.GradeSetting(
        year=item.year,
        grade=item.grade,
        class_count=item.class_count,
        required_homerooms=item.required_homerooms,
        required_subject_teachers=subject_teachers,
        required_duty_heads=duty_heads,
      )
      session.add(gs)
  await session.commit()
  return {"status": "ok"}


@router.get("/preferences/status")
async def get_preference_status(
  year: int,
  session: AsyncSession = Depends(get_session),
  user=Depends(get_current_user),
):
  """희망 제출 마감 상태 조회 (교사/관리자 모두 사용 가능)"""
  stmt = select(models.AdminSetting).where(models.AdminSetting.year == year)
  res = await session.execute(stmt)
  admin_setting = res.scalars().first()
  is_closed = admin_setting.is_closed if admin_setting else False
  return {"year": year, "is_closed": is_closed}


@router.put("/preferences/close")
async def close_preferences(
  payload: ClosePreferenceRequest,
  session: AsyncSession = Depends(get_session),
  user=Depends(get_current_user),
):
  """희망 제출 마감 설정/해제 (관리자만)"""
  if user.get("role") != "admin":
    raise HTTPException(status_code=403, detail="admin only")
  
  stmt = select(models.AdminSetting).where(models.AdminSetting.year == payload.year)
  res = await session.execute(stmt)
  admin_setting = res.scalars().first()
  
  if admin_setting:
    admin_setting.is_closed = payload.is_closed
  else:
    admin_setting = models.AdminSetting(
      year=payload.year,
      total_teachers=0,
      is_closed=payload.is_closed,
    )
    session.add(admin_setting)
  
  await session.commit()
  return {"year": payload.year, "is_closed": payload.is_closed}


@router.put("/dashboard/total-teachers")
async def update_total_teachers(
  payload: AdminSettingIn,
  session: AsyncSession = Depends(get_session),
  user=Depends(get_current_user),
):
  if user.get("role") != "admin":
    raise HTTPException(status_code=403, detail="admin only")
  
  stmt = select(models.AdminSetting).where(models.AdminSetting.year == payload.year)
  res = await session.execute(stmt)
  admin_setting = res.scalars().first()
  
  if admin_setting:
    admin_setting.total_teachers = payload.total_teachers
  else:
    admin_setting = models.AdminSetting(year=payload.year, total_teachers=payload.total_teachers, is_closed=False)
    session.add(admin_setting)
  
  await session.commit()
  await session.refresh(admin_setting)
  return AdminSettingOut.model_validate(admin_setting)


@router.post("/upload-teachers")
async def upload_teachers(
  file: UploadFile = File(...),
  session: AsyncSession = Depends(get_session),
  user=Depends(get_current_user),
):
  if user.get("role") != "admin":
    raise HTTPException(status_code=403, detail="admin only")
  
  if not file.filename.endswith((".xlsx", ".xls")):
    raise HTTPException(status_code=400, detail="엑셀 파일(.xlsx, .xls)만 업로드 가능합니다.")
  
  try:
    # 파일 읽기
    contents = await file.read()
    wb = load_workbook(io.BytesIO(contents))
    ws = wb.active
    
    # 헤더 행 찾기 (첫 번째 행)
    headers = [cell.value for cell in ws[1]]
    
    # 헤더 매핑 (다양한 형식 지원)
    header_map = {}
    for idx, header in enumerate(headers):
      if header is None:
        continue
      header_str = str(header).strip().lower()
      if "이름" in header_str or "name" in header_str:
        header_map["name"] = idx
      elif "이메일" in header_str or "email" in header_str:
        header_map["email"] = idx
      elif "구글" in header_str or "google" in header_str:
        header_map["google_id"] = idx
      elif "성별" in header_str or "gender" in header_str:
        header_map["gender"] = idx
      elif "총 경력" in header_str or "발령" in header_str or "hire" in header_str:
        header_map["hire_year"] = idx
      elif "본교" in header_str or "근무 시작" in header_str or "school_join" in header_str:
        header_map["school_join_year"] = idx
      elif "올해 학년" in header_str or "current_grade" in header_str:
        header_map["current_grade"] = idx
      elif "올해 학급" in header_str or "current_class" in header_str:
        header_map["current_class"] = idx
      elif "담임" in header_str and "올해" in header_str or "is_homeroom" in header_str:
        header_map["is_homeroom_current"] = idx
      elif "교과전담" in header_str or "is_subject" in header_str:
        header_map["is_subject_teacher"] = idx
      elif "업무부장" in header_str or "duty_role" in header_str:
        header_map["duty_role"] = idx
      elif "교과" in header_str or "subject" in header_str:
        header_map["subject"] = idx
      elif "특수" in header_str or "special" in header_str:
        header_map["special_conditions"] = idx
      elif "본교 담임 이력" in header_str or "학년이력" in header_str or "grade_history" in header_str or "담임 이력" in header_str:
        header_map["grade_history"] = idx
    
    if "name" not in header_map:
      raise HTTPException(status_code=400, detail="엑셀 파일에 '이름' 열이 없습니다.")
    
    # 데이터 행 처리
    success_count = 0
    error_count = 0
    errors: List[str] = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
      # 빈 행 건너뛰기
      if not row[header_map["name"]].value:
        continue
      
      try:
        name = str(row[header_map["name"]].value).strip()
        if not name:
          continue
        
        # 기존 교사 찾기 또는 새로 생성
        stmt = select(models.Teacher).where(models.Teacher.name == name)
        res = await session.execute(stmt)
        teacher = res.scalars().first()
        
        if not teacher:
          teacher = models.Teacher(name=name)
          session.add(teacher)
        
        # 데이터 업데이트
        if "email" in header_map and row[header_map["email"]].value:
          teacher.email = str(row[header_map["email"]].value).strip() or None
        
        if "google_id" in header_map and row[header_map["google_id"]].value:
          teacher.google_id = str(row[header_map["google_id"]].value).strip() or None
        
        if "gender" in header_map and row[header_map["gender"]].value:
          teacher.gender = str(row[header_map["gender"]].value).strip()[:10] or None
        
        if "hire_year" in header_map and row[header_map["hire_year"]].value:
          try:
            teacher.hire_year = int(row[header_map["hire_year"]].value)
          except (ValueError, TypeError):
            pass
        
        if "school_join_year" in header_map and row[header_map["school_join_year"]].value:
          try:
            teacher.school_join_year = int(row[header_map["school_join_year"]].value)
          except (ValueError, TypeError):
            pass
        
        if "current_grade" in header_map and row[header_map["current_grade"]].value:
          try:
            teacher.current_grade = int(row[header_map["current_grade"]].value)
          except (ValueError, TypeError):
            pass
        
        if "current_class" in header_map and row[header_map["current_class"]].value:
          teacher.current_class = str(row[header_map["current_class"]].value).strip() or None
        
        if "is_homeroom_current" in header_map:
          val = row[header_map["is_homeroom_current"]].value
          if val is not None:
            if isinstance(val, bool):
              teacher.is_homeroom_current = val
            elif isinstance(val, str):
              teacher.is_homeroom_current = val.lower() in ("true", "1", "예", "yes", "담임", "o", "○")
            elif isinstance(val, (int, float)):
              teacher.is_homeroom_current = bool(val)
        
        if "is_subject_teacher" in header_map:
          val = row[header_map["is_subject_teacher"]].value
          if val is not None:
            if isinstance(val, bool):
              teacher.is_subject_teacher = val
            elif isinstance(val, str):
              teacher.is_subject_teacher = val.lower() in ("true", "1", "예", "yes", "교과전담", "o", "○")
            elif isinstance(val, (int, float)):
              teacher.is_subject_teacher = bool(val)
        
        if "duty_role" in header_map and row[header_map["duty_role"]].value:
          teacher.duty_role = str(row[header_map["duty_role"]].value).strip() or None
        
        if "subject" in header_map and row[header_map["subject"]].value:
          teacher.subject = str(row[header_map["subject"]].value).strip() or None
        
        if "special_conditions" in header_map and row[header_map["special_conditions"]].value:
          teacher.special_conditions = str(row[header_map["special_conditions"]].value).strip() or None
        
        # 학년 이력 파싱 및 저장
        if "grade_history" in header_map and row[header_map["grade_history"]].value:
          import json
          history_str = str(row[header_map["grade_history"]].value).strip()
          if history_str:
            try:
              # "연도:학년,연도:학년" 형식 파싱
              history = []
              for entry in history_str.split(","):
                entry = entry.strip()
                if ":" in entry:
                  parts = entry.split(":")
                  if len(parts) == 2:
                    year = int(parts[0].strip())
                    grade = int(parts[1].strip())
                    history.append({"year": year, "grade": grade})
              if history:
                teacher.grade_history = json.dumps(history, ensure_ascii=False)
            except (ValueError, TypeError) as e:
              # 파싱 실패 시 무시 (에러 로그에 기록하지 않음)
              pass
        
        success_count += 1
        
      except Exception as e:
        error_count += 1
        errors.append(f"{row_idx}행: {str(e)}")
    
    await session.commit()
    
    return {
      "status": "ok",
      "success_count": success_count,
      "error_count": error_count,
      "errors": errors[:10],  # 최대 10개 오류만 반환
    }
    
  except Exception as e:
    raise HTTPException(status_code=400, detail=f"엑셀 파일 처리 중 오류: {str(e)}")


@router.post("/assign", response_model=list[AssignmentOut])
async def assign(
  year: int,
  session: AsyncSession = Depends(get_session),
  user=Depends(get_current_user),
):
  if user.get("role") != "admin":
    raise HTTPException(status_code=403, detail="admin only")
  try:
    # 기존 결과 삭제
    await session.execute(
      models.Assignment.__table__.delete().where(models.Assignment.year == year)
    )
    await session.commit()
    assigned, excluded, logs = await run_assignment(session, year)
    res = (
      await session.execute(select(models.Assignment).where(models.Assignment.year == year))
    ).scalars().all()
    return res
  except ValueError as e:
    raise HTTPException(status_code=400, detail=str(e))
  except Exception as e:
    raise HTTPException(status_code=500, detail=f"배정 중 오류 발생: {str(e)}")


@router.get("/assignments")
async def list_assignments(
  year: int,
  session: AsyncSession = Depends(get_session),
  user=Depends(get_current_user),
):
  if user.get("role") != "admin":
    raise HTTPException(status_code=403, detail="admin only")
  stmt = (
    select(
      models.Assignment.id,
      models.Assignment.teacher_id,
      models.Assignment.assigned_grade,
      models.Assignment.assignment_type,
      models.Assignment.rule_reference,
      models.Assignment.description,
      models.Teacher.name,
    )
    .join(models.Teacher, models.Teacher.id == models.Assignment.teacher_id)
    .where(models.Assignment.year == year)
  )
  res = await session.execute(stmt)
  rows = res.all()
  return [
    {
      "id": r.id,
      "teacher_id": r.teacher_id,
      "teacher_name": r.name,
      "assigned_grade": r.assigned_grade,
      "assignment_type": r.assignment_type,
      "rule_reference": r.rule_reference,
      "description": r.description,
    }
    for r in rows
  ]


@router.get("/assignments/export")
async def export_assignments(
  year: int,
  format: str = "csv",
  session: AsyncSession = Depends(get_session),
  user=Depends(get_current_user),
):
  if user.get("role") != "admin":
    raise HTTPException(status_code=403, detail="admin only")
  stmt = (
    select(
      models.Assignment.teacher_id,
      models.Teacher.name,
      models.Assignment.assigned_grade,
      models.Assignment.assignment_type,
      models.Assignment.rule_reference,
      models.Assignment.description,
    )
    .join(models.Teacher, models.Teacher.id == models.Assignment.teacher_id)
    .where(models.Assignment.year == year)
  )
  res = await session.execute(stmt)
  rows = res.all()

  output = io.StringIO()
  headers = ["teacher_id", "teacher_name", "assigned_grade", "assignment_type", "rule_reference", "description"]
  output.write(",".join(headers) + "\n")
  for r in rows:
    output.write(
      f"{r.teacher_id},{r.name},{r.assigned_grade},{r.assignment_type},{r.rule_reference or ''},{r.description or ''}\n"
    )
  output.seek(0)

  return StreamingResponse(
    output,
    media_type="text/csv",
    headers={"Content-Disposition": f'attachment; filename="assignments_{year}.csv"'},
  )


@router.get("/preferences")
async def list_preferences(
  year: int,
  session: AsyncSession = Depends(get_session),
  user=Depends(get_current_user),
):
  """제출한 사람들의 명단과 지망 정보 조회"""
  if user.get("role") != "admin":
    raise HTTPException(status_code=403, detail="admin only")
  
  import logging
  logger = logging.getLogger(__name__)
  logger.info(f"제출 명단 조회 요청: year={year}")
  
  # 전체 Preference 확인
  all_prefs_stmt = select(models.Preference).where(models.Preference.year == year)
  all_prefs_res = await session.execute(all_prefs_stmt)
  all_prefs = all_prefs_res.scalars().all()
  logger.info(f"해당 연도의 전체 Preference 수: {len(all_prefs)}")
  
  stmt = (
    select(
      models.Preference.id,
      models.Preference.teacher_id,
      models.Preference.first_choice_grade,
      models.Preference.second_choice_grade,
      models.Preference.third_choice_grade,
      models.Preference.wants_grade_head,
      models.Preference.wants_subject_teacher,
      models.Preference.wants_duty_head,
      models.Teacher.name,
    )
    .join(models.Teacher, models.Teacher.id == models.Preference.teacher_id)
    .where(models.Preference.year == year)
    .order_by(models.Teacher.name)
  )
  res = await session.execute(stmt)
  rows = res.all()
  
  result = [
    {
      "id": r.id,
      "teacher_id": r.teacher_id,
      "teacher_name": r.name,
      "first_choice_grade": r.first_choice_grade,
      "second_choice_grade": r.second_choice_grade,
      "third_choice_grade": r.third_choice_grade,
      "wants_grade_head": r.wants_grade_head,
      "wants_subject_teacher": r.wants_subject_teacher,
      "wants_duty_head": r.wants_duty_head,
    }
    for r in rows
  ]
  
  logger.info(f"제출 명단 조회 결과: {len(result)}건")
  return result


@router.delete("/preferences")
async def clear_preferences(
  year: int,
  session: AsyncSession = Depends(get_session),
  user=Depends(get_current_user),
):
  """특정 연도의 모든 희망 초기화"""
  if user.get("role") != "admin":
    raise HTTPException(status_code=403, detail="admin only")
  
  # 해당 연도의 모든 희망 삭제
  deleted = await session.execute(
    models.Preference.__table__.delete().where(models.Preference.year == year)
  )
  await session.commit()
  
  return {"status": "ok", "message": f"{year}년도 희망서가 모두 초기화되었습니다.", "deleted_count": deleted.rowcount}

