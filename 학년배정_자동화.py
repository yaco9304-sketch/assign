#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
학년 배정 자동화 스크립트
엑셀 파일을 읽어서 자동으로 학년을 배정하고 결과를 엑셀에 출력합니다.
"""

import json
from typing import List, Dict, Tuple, Optional
from dataclasses import dataclass, field
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import sys
import os


# ==================== 데이터 모델 ====================

@dataclass
class Teacher:
    """교사 정보"""
    name: str
    gender: Optional[str] = None
    hire_year: Optional[int] = None
    school_join_year: Optional[int] = None
    current_grade: Optional[int] = None
    current_class: Optional[str] = None
    is_homeroom_current: bool = False
    is_subject_teacher: bool = False
    duty_role: Optional[str] = None
    subject: Optional[str] = None
    special_conditions: Optional[str] = None
    grade_history: List[Dict[str, int]] = field(default_factory=list)
    banned_grades: set = field(default_factory=set)
    preferred_grade_primary: Optional[int] = None
    preferred_grade_secondary: Optional[int] = None
    preferred_grade_third: Optional[int] = None


@dataclass
class Preference:
    """희망사항"""
    teacher_name: str
    year: int
    first_choice_grade: Optional[int] = None
    second_choice_grade: Optional[int] = None
    third_choice_grade: Optional[int] = None
    wants_grade_head: bool = False
    wants_subject_teacher: bool = False
    wants_duty_head: bool = False
    comment: Optional[str] = None


@dataclass
class GradeSetting:
    """학년별 설정"""
    year: int
    grade: int
    class_count: int
    required_homerooms: int
    required_subject_teachers: int = 0
    required_duty_heads: int = 0


@dataclass
class Assignment:
    """배정 결과"""
    teacher_name: str
    assigned_grade: int
    assignment_type: str
    rule_reference: Optional[str] = None
    description: Optional[str] = None


# ==================== 배정 규칙 ====================

EXCLUDE_PATTERNS = [
    ("휴직", "제13조: 휴직"),
    ("병가", "제13조: 병가 30일 이상"),
    ("파견", "제13조: 파견"),
    ("연수", "제13조: 연수"),
    ("산전", "제13조: 임신/산전"),
    ("임신", "제13조: 임신"),
    ("출산", "제13조: 출산 예정"),
    ("고령", "제13조: 고령 교사"),
]

ROLE_POINTS = {
    "업무1부장": 6.0,
    "업무2부장": 5.0,
    "업무3부장": 4.3,
    "학년부장": 2.0,
    "교과전담": 3.0,
}

PRIORITY_PATTERNS = [
    ("원로", "제12조④: 원로교사"),
    ("요양", "제12조④: 요양 필요"),
    ("건강", "제12조④: 건강 사유"),
    ("군입대", "제12조③/제14조③: 군 입대"),
    ("출산", "제12조④: 출산 예정"),
    ("임신", "제12조④: 임신"),
]


def apply_exclusions(teachers: List[Teacher], year: int) -> Tuple[List[Teacher], List[Teacher], List[Tuple[str, str]]]:
    """제외 대상 필터링 (제13조)"""
    kept, excluded, logs = [], [], []
    for t in teachers:
        cond = (t.special_conditions or "").lower()
        reason = None
        for key, msg in EXCLUDE_PATTERNS:
            if key in cond:
                reason = msg
                break
        if reason:
            excluded.append(t)
            logs.append((t.name, "exclude", reason))
        else:
            kept.append(t)
    return kept, excluded, logs


def apply_priority_rules(teachers: List[Teacher], settings: List[GradeSetting], year: int, prefs_by_name: Dict[str, Preference]) -> Tuple[List[Tuple[Teacher, int, str, str]], List[Teacher], List[Tuple[str, str]]]:
    """우선 배정 규칙 적용"""
    assigned: List[Tuple[Teacher, int, str, str]] = []
    remaining: List[Teacher] = teachers[:]
    logs = []

    # 1) 특수 사유 우선 (제12조④)
    pri_candidates = []
    still = []
    for t in remaining:
        reason = None
        cond = (t.special_conditions or "").lower()
        for key, msg in PRIORITY_PATTERNS:
            if key in cond:
                reason = msg
                break
        if reason:
            pri_candidates.append((t, reason))
        else:
            still.append(t)
    remaining = still

    # 특수 사유자는 1지망→2지망→3지망 순으로 우선 배정
    for t, reason in pri_candidates:
        pref = prefs_by_name.get(t.name)
        prefs = []
        if pref:
            if pref.first_choice_grade:
                prefs.append(pref.first_choice_grade)
            if pref.second_choice_grade:
                prefs.append(pref.second_choice_grade)
            if pref.third_choice_grade:
                prefs.append(pref.third_choice_grade)
        
        target_grade = t.current_grade
        hope_detail = ""
        if not target_grade and prefs:
            target_grade = prefs[0]
            hope_detail = " (1지망 반영)"
        elif target_grade:
            hope_detail = " (현재 학년 유지)"
        
        if target_grade:
            desc = f"{reason}{hope_detail}"
            assigned.append((t, target_grade, "규정우선", desc))
            logs.append((t.name, "rule_12_4", reason))
        else:
            remaining.append(t)

    # 2) 업무부장/학년부장/교과전담 경합 시 점수 높은 순 (제12조②)
    def role_score(t: Teacher) -> float:
        if t.duty_role:
            for k, v in ROLE_POINTS.items():
                if k in t.duty_role:
                    return v
        return 0.0

    role_sorted = sorted(
        [t for t in remaining if role_score(t) > 0],
        key=lambda x: role_score(x),
        reverse=True,
    )
    remaining = [t for t in remaining if role_score(t) == 0]

    for t in role_sorted:
        pref = prefs_by_name.get(t.name)
        prefs = []
        if pref:
            if pref.first_choice_grade:
                prefs.append(pref.first_choice_grade)
            if pref.second_choice_grade:
                prefs.append(pref.second_choice_grade)
            if pref.third_choice_grade:
                prefs.append(pref.third_choice_grade)
        
        target = prefs[0] if prefs else (t.current_grade or None)
        hope_detail = ""
        if target:
            if prefs and target == prefs[0]:
                hope_detail = " (1지망 반영)"
            elif len(prefs) > 1 and target == prefs[1]:
                hope_detail = " (2지망 반영)"
            elif len(prefs) > 2 and target == prefs[2]:
                hope_detail = " (3지망 반영)"
            elif target == t.current_grade:
                hope_detail = " (현재 학년 유지)"
            
            role_name = t.duty_role or "역할"
            desc = f"제12조② 역할 우선 ({role_name}){hope_detail}"
            assigned.append((t, target, "규정우선", desc))
            logs.append((t.name, "rule_12_2", f"역할 {t.duty_role or ''} 우선 배정"))
        else:
            remaining.append(t)

    return assigned, remaining, logs


def apply_rotation(teachers: List[Teacher], prefs_by_name: Dict[str, Preference]) -> List[Teacher]:
    """학년 순환 규칙 적용"""
    updated: List[Teacher] = []
    for t in teachers:
        banned = set()
        
        # 1. 올해 담당 학년 제외 (제12조①)
        if t.current_grade:
            pref = prefs_by_name.get(t.name)
            wants_same = False
            if pref:
                wants_same = (
                    pref.first_choice_grade == t.current_grade
                    or pref.second_choice_grade == t.current_grade
                    or pref.third_choice_grade == t.current_grade
                )
            # 1학년/6학년은 동일학년 희망 시 제한 완화
            if not (t.current_grade in {1, 6} and wants_same):
                banned.add(t.current_grade)
        
        # 2. 동일 학년 2번 제한 (본교 근무 기간 동안)
        if t.grade_history:
            grade_counts: Dict[int, int] = {}
            for entry in t.grade_history:
                if isinstance(entry, dict) and "grade" in entry:
                    grade = entry["grade"]
                    grade_counts[grade] = grade_counts.get(grade, 0) + 1
            
            # 2번 이상 담임한 학년은 제외
            for grade, count in grade_counts.items():
                if count >= 2:
                    banned.add(grade)
        
        t.banned_grades = banned
        updated.append(t)
    return updated


def apply_subject_rules(teachers: List[Teacher]) -> List[Teacher]:
    """교과전담 규칙 적용"""
    updated: List[Teacher] = []
    for t in teachers:
        if t.is_subject_teacher:
            banned = getattr(t, "banned_grades", set())
            banned.update({1, 2, 3, 4, 5, 6})
            t.banned_grades = banned
        updated.append(t)
    return updated


def score_candidate(teacher: Teacher, grade: int, prefs: List[int]) -> Tuple[int, Dict]:
    """점수 계산 및 상세 내역 반환"""
    hope_score = 0
    hope_detail = ""
    if prefs:
        if grade == prefs[0]:
            hope_score = 10
            hope_detail = "1지망"
        elif len(prefs) > 1 and grade == prefs[1]:
            hope_score = 5
            hope_detail = "2지망"
        elif len(prefs) > 2 and grade == prefs[2]:
            hope_score = 2
            hope_detail = "3지망"
    
    grade_weight = {6: 6, 1: 5, 5: 4, 3: 3, 4: 3, 2: 2}.get(grade, 0)
    
    role_score = 0
    role_detail = ""
    role_text = (teacher.duty_role or "") + " " + (teacher.special_conditions or "") + " " + (teacher.subject or "")
    for key, val in ROLE_POINTS.items():
        if key in role_text:
            if val > role_score:
                role_score = val
                role_detail = key
    
    penalty = -999 if grade in teacher.banned_grades else 0
    total_score = hope_score + grade_weight + role_score + penalty
    
    details = {
        "hope_score": hope_score,
        "hope_detail": hope_detail,
        "grade_weight": grade_weight,
        "role_score": role_score,
        "role_detail": role_detail,
        "penalty": penalty,
        "total_score": total_score,
    }
    
    return total_score, details


def run_assignment(teachers: List[Teacher], settings: List[GradeSetting], prefs_by_name: Dict[str, Preference], year: int) -> Tuple[List[Assignment], List[Teacher], List[Tuple[str, str]]]:
    """배정 알고리즘 실행"""
    # 1. 제외 대상 필터링
    kept, excluded, logs_all = apply_exclusions(teachers, year)
    
    # 2. 희망 정보를 teacher 객체에 붙이기
    for t in kept:
        pref = prefs_by_name.get(t.name)
        if pref:
            t.preferred_grade_primary = pref.first_choice_grade
            t.preferred_grade_secondary = pref.second_choice_grade
            t.preferred_grade_third = pref.third_choice_grade
    
    # 3. 우선 배정 규칙 적용
    assigned, remaining, pri_logs = apply_priority_rules(kept, settings, year, prefs_by_name)
    logs_all.extend(pri_logs)
    
    # 4. 학년 순환 규칙 적용
    remaining = apply_rotation(remaining, prefs_by_name)
    
    # 5. 교과전담 규칙 적용
    remaining = apply_subject_rules(remaining)
    
    # 6. 슬롯 풀 생성
    slots: List[int] = []
    for s in settings:
        for _ in range(s.required_homerooms):
            slots.append(s.grade)
    
    if not slots:
        raise ValueError("필요 담임 수가 0입니다. 학급 설정에서 필요 담임 수를 입력해주세요.")
    
    # 7. 1/2/3 지망 우선 배정
    for choice_idx in [0, 1, 2]:
        still = []
        for t in remaining:
            prefs = []
            pref = prefs_by_name.get(t.name)
            if pref:
                if pref.first_choice_grade:
                    prefs.append(pref.first_choice_grade)
                if pref.second_choice_grade:
                    prefs.append(pref.second_choice_grade)
                if pref.third_choice_grade:
                    prefs.append(pref.third_choice_grade)
            if choice_idx < len(prefs) and prefs[choice_idx] in slots and prefs[choice_idx] not in t.banned_grades:
                g = prefs[choice_idx]
                hope_rank = f"{choice_idx+1}지망"
                desc = f"{hope_rank} 반영 (희망 학년: {g}학년)"
                assigned.append((t, g, hope_rank, desc))
                slots.remove(g)
            else:
                still.append(t)
        remaining = still
    
    # 8. 남은 슬롯 점수 기반 배정 (greedy)
    scored = []
    for t in remaining:
        prefs = []
        pref = prefs_by_name.get(t.name)
        if pref:
            if pref.first_choice_grade:
                prefs.append(pref.first_choice_grade)
            if pref.second_choice_grade:
                prefs.append(pref.second_choice_grade)
            if pref.third_choice_grade:
                prefs.append(pref.third_choice_grade)
        best = None
        best_details = None
        for g in set(slots):
            sc, details = score_candidate(t, g, prefs)
            if best is None or sc > best[1]:
                best = (g, sc)
                best_details = details
        if best:
            scored.append((t, best[0], best[1], best_details))
    scored.sort(key=lambda x: x[2], reverse=True)
    for t, g, sc, details in scored:
        if g in slots:
            desc_parts = []
            if details["hope_detail"]:
                desc_parts.append(f"희망: {details['hope_detail']}({details['hope_score']}점)")
            desc_parts.append(f"학년가중치: {details['grade_weight']}점")
            if details["role_detail"]:
                desc_parts.append(f"역할: {details['role_detail']}({details['role_score']}점)")
            desc_parts.append(f"총점: {details['total_score']}점")
            desc = " | ".join(desc_parts)
            assigned.append((t, g, "조정", desc))
            slots.remove(g)
    
    # 9. 결과 변환
    assignments = []
    for t, g, atype, desc in assigned:
        rule_ref = None
        if "규정우선" in atype:
            if "제12조④" in (desc or ""):
                rule_ref = "제12조④ (특수 사유 우선 배정)"
            elif "제12조②" in (desc or ""):
                rule_ref = "제12조② (역할 우선 배정)"
            elif "제13조" in (desc or ""):
                rule_ref = "제13조 (배정 제외)"
        elif "1지망" in atype or "2지망" in atype or "3지망" in atype:
            rule_ref = "제11조 (희망 학년 반영)"
        elif "조정" in atype:
            rule_ref = "제12조① (학년 순환 원칙) + 점수 기반 조정"
        
        assignments.append(Assignment(
            teacher_name=t.name,
            assigned_grade=g,
            assignment_type=atype,
            rule_reference=rule_ref,
            description=desc,
        ))
    
    return assignments, excluded, logs_all


# ==================== 엑셀 읽기/쓰기 ====================

def read_teachers_from_excel(wb: Workbook) -> List[Teacher]:
    """교사 정보 시트 읽기"""
    if "교사정보" not in wb.sheetnames:
        raise ValueError("'교사정보' 시트가 없습니다.")
    
    ws = wb["교사정보"]
    teachers = []
    
    # 헤더 찾기
    header_row = 1
    headers = {}
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(header_row, col).value
        if cell_value:
            headers[str(cell_value).strip()] = col
    
    # 데이터 읽기
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, headers.get("이름", 1)).value
        if not name:
            continue
        
        # 본교 담임 이력 파싱
        grade_history = []
        history_str = ws.cell(row, headers.get("본교 담임 이력", 0)).value
        if history_str:
            try:
                # 형식: "2023:1,2024:2" 또는 JSON
                if history_str.startswith("["):
                    grade_history = json.loads(history_str)
                else:
                    for pair in str(history_str).split(","):
                        if ":" in pair:
                            year_str, grade_str = pair.split(":")
                            grade_history.append({"year": int(year_str.strip()), "grade": int(grade_str.strip())})
            except:
                pass
        
        teacher = Teacher(
            name=str(name).strip(),
            gender=str(ws.cell(row, headers.get("성별", 0)).value or "").strip() or None,
            hire_year=int(ws.cell(row, headers.get("임용년도", 0)).value) if ws.cell(row, headers.get("임용년도", 0)).value else None,
            school_join_year=int(ws.cell(row, headers.get("본교입사년도", 0)).value) if ws.cell(row, headers.get("본교입사년도", 0)).value else None,
            current_grade=int(ws.cell(row, headers.get("현재 학년", 0)).value) if ws.cell(row, headers.get("현재 학년", 0)).value else None,
            current_class=str(ws.cell(row, headers.get("현재 반", 0)).value or "").strip() or None,
            is_homeroom_current=bool(ws.cell(row, headers.get("현재 담임 여부", 0)).value) if headers.get("현재 담임 여부", 0) else False,
            is_subject_teacher=bool(ws.cell(row, headers.get("교과전담 여부", 0)).value) if headers.get("교과전담 여부", 0) else False,
            duty_role=str(ws.cell(row, headers.get("역할", 0)).value or "").strip() or None,
            subject=str(ws.cell(row, headers.get("담당 교과", 0)).value or "").strip() or None,
            special_conditions=str(ws.cell(row, headers.get("특수 조건", 0)).value or "").strip() or None,
            grade_history=grade_history,
        )
        teachers.append(teacher)
    
    return teachers


def read_preferences_from_excel(wb: Workbook, year: int) -> Dict[str, Preference]:
    """희망사항 시트 읽기"""
    if "희망사항" not in wb.sheetnames:
        return {}
    
    ws = wb["희망사항"]
    prefs_by_name = {}
    
    # 헤더 찾기
    header_row = 1
    headers = {}
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(header_row, col).value
        if cell_value:
            headers[str(cell_value).strip()] = col
    
    # 데이터 읽기
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, headers.get("교사 이름", 1)).value
        if not name:
            continue
        
        pref = Preference(
            teacher_name=str(name).strip(),
            year=year,
            first_choice_grade=int(ws.cell(row, headers.get("1지망", 0)).value) if ws.cell(row, headers.get("1지망", 0)).value else None,
            second_choice_grade=int(ws.cell(row, headers.get("2지망", 0)).value) if ws.cell(row, headers.get("2지망", 0)).value else None,
            third_choice_grade=int(ws.cell(row, headers.get("3지망", 0)).value) if ws.cell(row, headers.get("3지망", 0)).value else None,
            wants_grade_head=bool(ws.cell(row, headers.get("학년부장", 0)).value) if headers.get("학년부장", 0) else False,
            wants_subject_teacher=bool(ws.cell(row, headers.get("교과전담", 0)).value) if headers.get("교과전담", 0) else False,
            wants_duty_head=bool(ws.cell(row, headers.get("업무부장", 0)).value) if headers.get("업무부장", 0) else False,
            comment=str(ws.cell(row, headers.get("비고", 0)).value or "").strip() or None,
        )
        prefs_by_name[pref.teacher_name] = pref
    
    return prefs_by_name


def read_grade_settings_from_excel(wb: Workbook, year: int) -> List[GradeSetting]:
    """학년 설정 시트 읽기"""
    if "학년설정" not in wb.sheetnames:
        raise ValueError("'학년설정' 시트가 없습니다.")
    
    ws = wb["학년설정"]
    settings = []
    
    # 헤더 찾기
    header_row = 1
    headers = {}
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(header_row, col).value
        if cell_value:
            headers[str(cell_value).strip()] = col
    
    # 데이터 읽기
    for row in range(2, ws.max_row + 1):
        grade = ws.cell(row, headers.get("학년", 1)).value
        if not grade:
            continue
        
        setting = GradeSetting(
            year=year,
            grade=int(grade),
            class_count=int(ws.cell(row, headers.get("학급 수", 0)).value) if ws.cell(row, headers.get("학급 수", 0)).value else 0,
            required_homerooms=int(ws.cell(row, headers.get("필요 담임 수", 0)).value) if ws.cell(row, headers.get("필요 담임 수", 0)).value else 0,
            required_subject_teachers=int(ws.cell(row, headers.get("필요 교과전담 수", 0)).value) if ws.cell(row, headers.get("필요 교과전담 수", 0)).value else 0,
            required_duty_heads=int(ws.cell(row, headers.get("필요 업무부장 수", 0)).value) if ws.cell(row, headers.get("필요 업무부장 수", 0)).value else 0,
        )
        settings.append(setting)
    
    return settings


def write_results_to_excel(wb: Workbook, assignments: List[Assignment], excluded: List[Teacher], year: int):
    """배정 결과를 엑셀에 쓰기"""
    # 기존 시트 삭제 (있다면)
    if "배정결과" in wb.sheetnames:
        wb.remove(wb["배정결과"])
    
    ws = wb.create_sheet("배정결과")
    
    # 헤더 스타일
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    
    # 헤더 작성
    headers = ["교사 이름", "배정 학년", "배정 유형", "적용 규정", "상세 근거"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # 배정 결과 작성
    for row, assignment in enumerate(assignments, 2):
        ws.cell(row, 1, assignment.teacher_name).border = border
        ws.cell(row, 2, assignment.assigned_grade).border = border
        ws.cell(row, 3, assignment.assignment_type).border = border
        ws.cell(row, 4, assignment.rule_reference or "").border = border
        ws.cell(row, 5, assignment.description or "").border = border
    
    # 제외 대상 작성
    if excluded:
        start_row = len(assignments) + 3
        ws.cell(start_row, 1, "제외 대상").font = Font(bold=True, size=12)
        ws.cell(start_row + 1, 1, "교사 이름").fill = header_fill
        ws.cell(start_row + 1, 1).font = header_font
        ws.cell(start_row + 1, 1).alignment = header_alignment
        ws.cell(start_row + 1, 1).border = border
        ws.cell(start_row + 1, 2, "제외 사유").fill = header_fill
        ws.cell(start_row + 1, 2).font = header_font
        ws.cell(start_row + 1, 2).alignment = header_alignment
        ws.cell(start_row + 1, 2).border = border
        
        for idx, teacher in enumerate(excluded, 2):
            ws.cell(start_row + idx, 1, teacher.name).border = border
            reason = teacher.special_conditions or "제13조: 배정 제외"
            ws.cell(start_row + idx, 2, reason).border = border
    
    # 열 너비 자동 조정
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    # 학년별 통계 시트 생성
    if "학년별통계" in wb.sheetnames:
        wb.remove(wb["학년별통계"])
    
    stats_ws = wb.create_sheet("학년별통계")
    stats_headers = ["학년", "배정 인원 수"]
    for col, header in enumerate(stats_headers, 1):
        cell = stats_ws.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # 학년별 통계 계산
    grade_counts: Dict[int, int] = {}
    for assignment in assignments:
        grade_counts[assignment.assigned_grade] = grade_counts.get(assignment.assigned_grade, 0) + 1
    
    for row, (grade, count) in enumerate(sorted(grade_counts.items()), 2):
        stats_ws.cell(row, 1, grade).border = border
        stats_ws.cell(row, 2, count).border = border
    
    for col in range(1, len(stats_headers) + 1):
        stats_ws.column_dimensions[get_column_letter(col)].width = 20


def add_execution_sheet(wb: Workbook, excel_path: str):
    """배정 실행 시트 추가 (버튼 포함)"""
    # 기존 시트 삭제 (있다면)
    if "배정실행" in wb.sheetnames:
        wb.remove(wb["배정실행"])
    
    ws = wb.create_sheet("배정실행", 0)  # 첫 번째 시트로 추가
    
    # 스타일 설정
    title_font = Font(bold=True, size=16, color="366092")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    button_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    button_font = Font(bold=True, color="FFFFFF", size=12)
    info_font = Font(size=10, color="666666")
    
    # 제목
    ws.cell(2, 2, "학년 배정 자동화 시스템").font = title_font
    ws.merge_cells("B2:F2")
    
    # 안내 문구
    ws.cell(4, 2, "📋 사용 방법:")
    ws.cell(5, 2, "1. 교사정보, 희망사항, 학년설정 시트에 데이터를 입력하세요.")
    ws.cell(6, 2, "2. 아래 '배정 실행' 버튼을 클릭하거나 VBA 매크로를 실행하세요.")
    ws.cell(7, 2, "3. 배정 결과는 '배정결과' 시트에 자동으로 생성됩니다.")
    
    for row in range(4, 8):
        ws.cell(row, 2).font = info_font
    
    # 배정 연도 입력
    ws.cell(10, 2, "배정 연도:")
    ws.cell(10, 2).font = Font(bold=True)
    year_cell = ws.cell(10, 3, datetime.now().year + 1)
    year_cell.border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    year_cell.alignment = Alignment(horizontal="center")
    
    # 버튼 영역 (시각적으로 버튼처럼 보이게)
    button_row = 12
    button_col = 2
    button_cell = ws.cell(button_row, button_col, "▶ 배정 실행")
    button_cell.fill = button_fill
    button_cell.font = button_font
    button_cell.alignment = Alignment(horizontal="center", vertical="center")
    button_cell.border = Border(
        left=Side(style="medium"),
        right=Side(style="medium"),
        top=Side(style="medium"),
        bottom=Side(style="medium"),
    )
    ws.merge_cells(f"B{button_row}:D{button_row}")
    ws.row_dimensions[button_row].height = 30
    
    # Python 스크립트 경로 저장 (숨김 셀)
    script_path = os.path.abspath(__file__)
    excel_dir = os.path.dirname(os.path.abspath(excel_path))
    relative_script = os.path.relpath(script_path, excel_dir)
    ws.cell(1, 1, f"SCRIPT_PATH:{relative_script}")  # 숨김 정보
    
    # VBA 매크로 사용 안내
    ws.cell(15, 2, "💡 VBA 매크로 설정 방법:")
    ws.cell(15, 2).font = Font(bold=True, size=11)
    ws.cell(16, 2, "1. Alt+F11을 눌러 VBA 편집기를 엽니다.")
    ws.cell(17, 2, "2. ThisWorkbook에 '엑셀_자동실행.bas' 파일의 코드를 붙여넣습니다.")
    ws.cell(18, 2, "3. 저장 후 '배정 실행' 버튼을 클릭하면 자동으로 배정이 실행됩니다.")
    
    for row in range(16, 19):
        ws.cell(row, 2).font = info_font
    
    # 열 너비 조정
    ws.column_dimensions["A"].width = 1
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15


def create_template_excel(filename: str):
    """템플릿 엑셀 파일 생성"""
    wb = Workbook()
    
    # 교사정보 시트
    ws_teachers = wb.active
    ws_teachers.title = "교사정보"
    headers_teachers = [
        "이름", "성별", "임용년도", "본교입사년도", "현재 학년", "현재 반",
        "현재 담임 여부", "교과전담 여부", "역할", "담당 교과", "특수 조건", "본교 담임 이력"
    ]
    for col, header in enumerate(headers_teachers, 1):
        cell = ws_teachers.cell(1, col, header)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, len(headers_teachers) + 1):
        ws_teachers.column_dimensions[get_column_letter(col)].width = 15
    
    # 희망사항 시트
    ws_prefs = wb.create_sheet("희망사항")
    headers_prefs = [
        "교사 이름", "1지망", "2지망", "3지망", "학년부장", "교과전담", "업무부장", "비고"
    ]
    for col, header in enumerate(headers_prefs, 1):
        cell = ws_prefs.cell(1, col, header)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, len(headers_prefs) + 1):
        ws_prefs.column_dimensions[get_column_letter(col)].width = 15
    
    # 학년설정 시트
    ws_settings = wb.create_sheet("학년설정")
    headers_settings = [
        "학년", "학급 수", "필요 담임 수", "필요 교과전담 수", "필요 업무부장 수"
    ]
    for col, header in enumerate(headers_settings, 1):
        cell = ws_settings.cell(1, col, header)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, len(headers_settings) + 1):
        ws_settings.column_dimensions[get_column_letter(col)].width = 15
    
    # 배정 실행 시트 추가
    add_execution_sheet(wb, filename)
    
    wb.save(filename)
    print(f"✅ 템플릿 파일 생성 완료: {filename}")


# ==================== 메인 함수 ====================

def process_excel_file(excel_path: str, year: Optional[int] = None):
    """엑셀 파일 처리 (함수로 분리하여 VBA에서 호출 가능)"""
    if year is None:
        year = datetime.now().year + 1
    
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {excel_path}")
    
    print(f"📖 엑셀 파일 읽는 중: {excel_path}")
    wb = load_workbook(excel_path)
    
    # 배정 실행 시트에서 연도 읽기 시도
    if "배정실행" in wb.sheetnames:
        try:
            ws_exec = wb["배정실행"]
            year_cell_value = ws_exec.cell(10, 3).value
            if year_cell_value and isinstance(year_cell_value, (int, float)):
                year = int(year_cell_value)
                print(f"📅 배정 실행 시트에서 연도 읽음: {year}")
        except:
            pass
    
    print("📋 교사 정보 읽는 중...")
    teachers = read_teachers_from_excel(wb)
    print(f"   - 총 {len(teachers)}명의 교사 정보를 읽었습니다.")
    
    print("📋 희망사항 읽는 중...")
    prefs_by_name = read_preferences_from_excel(wb, year)
    print(f"   - 총 {len(prefs_by_name)}명의 희망사항을 읽었습니다.")
    
    print("📋 학년 설정 읽는 중...")
    settings = read_grade_settings_from_excel(wb, year)
    print(f"   - 총 {len(settings)}개 학년의 설정을 읽었습니다.")
    
    print("🔄 배정 알고리즘 실행 중...")
    try:
        assignments, excluded, logs = run_assignment(teachers, settings, prefs_by_name, year)
        print(f"   - 배정 완료: {len(assignments)}명")
        print(f"   - 제외 대상: {len(excluded)}명")
    except Exception as e:
        print(f"❌ 배정 중 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        raise
    
    print("💾 결과를 엑셀에 저장 중...")
    write_results_to_excel(wb, assignments, excluded, year)
    
    # 원본 파일에 결과 저장 (배정결과 시트 추가)
    wb.save(excel_path)
    
    # 별도 결과 파일도 생성
    output_path = excel_path.replace(".xlsx", f"_배정결과_{year}.xlsx")
    wb.save(output_path)
    print(f"✅ 배정 완료! 결과 파일: {output_path}")
    print(f"✅ 원본 파일에도 결과가 저장되었습니다: {excel_path}")
    print("")
    print("📊 배정 결과 요약:")
    grade_counts: Dict[int, int] = {}
    for assignment in assignments:
        grade_counts[assignment.assigned_grade] = grade_counts.get(assignment.assigned_grade, 0) + 1
    for grade in sorted(grade_counts.keys()):
        print(f"   - {grade}학년: {grade_counts[grade]}명")
    
    return assignments, excluded


def main():
    """메인 실행 함수"""
    if len(sys.argv) < 2:
        print("사용법:")
        print("  python 학년배정_자동화.py <엑셀파일경로> [배정연도]")
        print("  python 학년배정_자동화.py --template <템플릿파일명>")
        print("")
        print("예시:")
        print("  python 학년배정_자동화.py --template 학년배정_템플릿.xlsx")
        print("  python 학년배정_자동화.py 학년배정_템플릿.xlsx 2026")
        sys.exit(1)
    
    if sys.argv[1] == "--template":
        if len(sys.argv) < 3:
            filename = "학년배정_템플릿.xlsx"
        else:
            filename = sys.argv[2]
        create_template_excel(filename)
        return
    
    excel_path = sys.argv[1]
    year = int(sys.argv[2]) if len(sys.argv) > 2 else None
    
    try:
        process_excel_file(excel_path, year)
    except Exception as e:
        print(f"❌ 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    
    if not os.path.exists(excel_path):
        print(f"❌ 파일을 찾을 수 없습니다: {excel_path}")
        sys.exit(1)
    
    print(f"📖 엑셀 파일 읽는 중: {excel_path}")
    wb = load_workbook(excel_path)
    
    print("📋 교사 정보 읽는 중...")
    teachers = read_teachers_from_excel(wb)
    print(f"   - 총 {len(teachers)}명의 교사 정보를 읽었습니다.")
    
    print("📋 희망사항 읽는 중...")
    prefs_by_name = read_preferences_from_excel(wb, year)
    print(f"   - 총 {len(prefs_by_name)}명의 희망사항을 읽었습니다.")
    
    print("📋 학년 설정 읽는 중...")
    settings = read_grade_settings_from_excel(wb, year)
    print(f"   - 총 {len(settings)}개 학년의 설정을 읽었습니다.")
    
    print("🔄 배정 알고리즘 실행 중...")
    try:
        assignments, excluded, logs = run_assignment(teachers, settings, prefs_by_name, year)
        print(f"   - 배정 완료: {len(assignments)}명")
        print(f"   - 제외 대상: {len(excluded)}명")
    except Exception as e:
        print(f"❌ 배정 중 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    
    print("💾 결과를 엑셀에 저장 중...")
    write_results_to_excel(wb, assignments, excluded, year)
    
    # 원본 파일에 결과 저장 (배정결과 시트 추가)
    wb.save(excel_path)
    
    # 별도 결과 파일도 생성
    output_path = excel_path.replace(".xlsx", f"_배정결과_{year}.xlsx")
    wb.save(output_path)
    print(f"✅ 배정 완료! 결과 파일: {output_path}")
    print(f"✅ 원본 파일에도 결과가 저장되었습니다: {excel_path}")
    print("")
    print("📊 배정 결과 요약:")
    grade_counts: Dict[int, int] = {}
    for assignment in assignments:
        grade_counts[assignment.assigned_grade] = grade_counts.get(assignment.assigned_grade, 0) + 1
    for grade in sorted(grade_counts.keys()):
        print(f"   - {grade}학년: {grade_counts[grade]}명")


if __name__ == "__main__":
    main()

